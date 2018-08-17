import sys
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Alignment, Border, Side

filename=sys.argv[1]
df=pd.read_excel(filename)
df=df.sort_values(by=["前年度ランク"], ascending=True)

rank=df[df["前年度ランク"]>=0]
temp=df[~(df["前年度ランク"]>=0)]
temp=temp.reset_index(drop=True)

K=np.zeros(temp.shape[0])
for i in range(K.shape[0]):
    K[i]=np.sum(df['チーム名']==temp['チーム名'][i])
    
temp['チーム内ランク']=temp['チーム内ランク']/K
temp=temp.sort_values(by=["チーム内ランク", "チームランク"], ascending=True)

rank=pd.concat([rank, temp])
rank=rank.reset_index(drop=True)
rank.index+=1
rank.to_excel('rank_'+filename)

#tournament pairs considering rank
N=rank.shape[0]

position=np.array([1, 4, 3, 2])
p1=2
while N>pow(2, p1):
    i=0
    while i<pow(2, p1-1):
        position=np.r_[position[:4*i+1], np.zeros(2), position[4*i+1:]]
        i+=1
        
    i=0
    while i<pow(2, p1):
        if position[2*i]==0:
            position[2*i]=pow(2, p1+1)+1-position[2*i+1]
        else:
            position[2*i+1]=pow(2, p1+1)+1-position[2*i]
        i+=1    
    p1+=1
    
p2=1
while N-pow(2, p1-1)>pow(2, p2):
    p2+=1
    
position=position[position<=pow(2, p1-1)+pow(2, p2)]
position[position>N]=0



wb=openpyxl.Workbook()
ws=wb.active
ws.title='tournament'

#ヘッダー部を2行ごとに結合
#前6列にID, 選手姓, 選手名, '(', チーム名, ')'
n_header=6
i=1
while i<=position.shape[0]:
    #i行目の処理
    
    #i行目j列の結合と設定
    j=1
    while j<=n_header:
        ws.merge_cells(start_row=2*i-1, start_column=j, end_row=2*i, end_column=j)
        ws.cell(row=2*i-1, column=j).alignment=Alignment(
            wrap_text=False,
            horizontal='center',
            vertical='center',
        )
        j+=1

    #ここで列幅の自動調整をさせたいができていない
    #while j<=n_header:
    #    ws.col_dimensions[j].alignment=Alignment(
    #        ShrinkToFit=True
    #    )
    #    j+=1
    
    #ヘッダー中身書き込み
    ws.cell(row=2*i-1, column=1, value=i)
    ws.cell(row=2*i-1, column=4, value='(')
    ws.cell(row=2*i-1, column=6, value=')')
    
    if position[i-1]==0:
        ws.cell(row=2*i-1, column=2, value='bye')
    else:
        ws.cell(row=2*i-1, column=2, value=rank['選手姓'][position[i-1]])
        ws.cell(row=2*i-1, column=3, value=rank['選手名'][position[i-1]])
        ws.cell(row=2*i-1, column=5, value=rank['チーム名'][position[i-1]])
    i+=1

#1, 2回戦分
n_round1=1
n_round2=1
list_index=[]
n_seed=pow(2, p2)
i=1
while i<position.shape[0]:
    temp=position[i-1:i+2]
    if np.min(temp)==0:
        temp[np.argmin(temp)]=np.max(temp)
        
    if np.min(temp)<=n_seed:
        #連続する3人の中にシードが含まれる場合
        if np.argmin(temp)==0:
            #若番側がシードの場合
            ws.cell(row=2*i, column=n_header+1).border=Border(
                outline=True,
                top=Side(style='medium', color='FF000000')
            )
            ws.cell(row=2*i, column=n_header+2).border=Border(
                outline=True,
                top=Side(style='medium', color='FF000000'),
                right=Side(style='medium', color='FF000000')
            )
            ws.cell(row=2*(i+1), column=n_header+1).border=Border(
                outline=True,
                top=Side(style='medium', color='FF000000'),
                right=Side(style='medium', color='FF000000')
            )
            ws.cell(row=2*(i+1)+1, column=n_header+1).border=Border(
                outline=True,
                bottom=Side(style='medium', color='FF000000'),
                right=Side(style='medium', color='FF000000')
            )
            ws.cell(row=2*(i+1), column=n_header+2).border=Border(
                outline=True,
                bottom=Side(style='medium', color='FF000000'),
                right=Side(style='medium', color='FF000000')
            )
            ws.cell(row=2*(i+1)-1, column=n_header+2).border=Border(
                outline=True,
                right=Side(style='medium', color='FF000000')
            )
            ws.cell(row=2*i, column=n_header+3).border=Border(
                outline=True,
                bottom=Side(style='medium', color='FF000000'),
            )
            ws.cell(row=2*i, column=n_header+2, value='2-'+str(n_round2))
            ws.cell(row=2*(i+1), column=n_header+1, value='1-'+str(n_round1))
            ws.cell(row=2*i, column=n_header+2).alignment=Alignment(
                wrap_text=False,
                horizontal='right',
                vertical='center',
            )
            ws.cell(row=2*(i+1), column=n_header+1).alignment=Alignment(
                wrap_text=False,
                horizontal='right',
                vertical='center',
            )
            n_round1+=1
            n_round2+=1
            list_index.append(2*i)
        else:
            #若番側は1回戦スタートの場合
            ws.cell(row=2*i, column=n_header+1).border=Border(
                outline=True,
                top=Side(style='medium', color='FF000000'),
                right=Side(style='medium', color='FF000000')
            )
            ws.cell(row=2*i+1, column=n_header+1).border=Border(
                outline=True,
                bottom=Side(style='medium', color='FF000000'),
                right=Side(style='medium', color='FF000000')
            )
            ws.cell(row=2*i+1, column=n_header+2).border=Border(
                outline=True,
                top=Side(style='medium', color='FF000000'),
                right=Side(style='medium', color='FF000000')
            )
            ws.cell(row=2*(i+1), column=n_header+2).border=Border(
                outline=True,
                right=Side(style='medium', color='FF000000')
            )
            ws.cell(row=2*i+3, column=n_header+1).border=Border(
                outline=True,
                bottom=Side(style='medium', color='FF000000')
            )
            ws.cell(row=2*i+3, column=n_header+2).border=Border(
                outline=True,
                bottom=Side(style='medium', color='FF000000'),
                right=Side(style='medium', color='FF000000')
            )
            ws.cell(row=2*(i+1), column=n_header+3).border=Border(
                outline=True,
                bottom=Side(style='medium', color='FF000000')
            )
            ws.cell(row=2*i, column=n_header+1, value='1-'+str(n_round1))
            ws.cell(row=2*(i+1), column=n_header+2, value='2-'+str(n_round2))
            ws.cell(row=2*i, column=n_header+1).alignment=Alignment(
                wrap_text=False,
                horizontal='right',
                vertical='center',
            )
            ws.cell(row=2*(i+1), column=n_header+2).alignment=Alignment(
                wrap_text=False,
                horizontal='right',
                vertical='center',
            )
            n_round1+=1
            n_round2+=1
            list_index.append(2*(i+1))
        i+=3
    else:
        #連続する3人にシードがいない場合
        ws.cell(row=2*i, column=n_header+1).border=Border(
            outline=True,
            top=Side(style='medium', color='FF000000')
        )
        ws.cell(row=2*i+1, column=n_header+1).border=Border(
            outline=True,
            bottom=Side(style='medium', color='FF000000')
        )
        ws.cell(row=2*i, column=n_header+2).border=Border(
            outline=True,
            top=Side(style='medium', color='FF000000'),
            right=Side(style='medium', color='FF000000')
        )
        ws.cell(row=2*i+1, column=n_header+2).border=Border(
            outline=True,
            bottom=Side(style='medium', color='FF000000'),
            right=Side(style='medium', color='FF000000')
        )
        ws.cell(row=2*i, column=n_header+3).border=Border(
            outline=True,
            bottom=Side(style='medium', color='FF000000')
        )
        ws.cell(row=2*i, column=n_header+2, value='2-'+str(n_round2))
        ws.cell(row=2*i, column=n_header+2).alignment=Alignment(
            wrap_text=False,
            horizontal='right',
            vertical='center',
        )
        n_round2+=1
        list_index.append(2*i)
        i+=2

#3回戦以降
r=3 #階層数
while len(list_index)>1:
    temp=[] #次のlist_index
    n_roundr=1
    i=0
    while i<len(list_index):
        x=list_index[i]+1
        ws.cell(row=x, column=n_header+r).border=Border(
            outline=True,
            top=Side('medium', color='FF000000'),
            right=Side(style='medium', color='FF000000')
        )
        x+=1
        while x<list_index[i+1]:
            ws.cell(row=x, column=n_header+r).border=Border(
                outline=True,
                right=Side(style='medium', color='FF000000')
            )
            x+=1
        ws.cell(row=x, column=n_header+r).border=Border(
            outline=True,
            bottom=Side('medium', color='FF000000'),
            right=Side(style='medium', color='FF000000')
        )
        
        ws.cell(row=(list_index[i]+list_index[i+1])/2, column=n_header+r, value=str(r)+'-'+str(n_roundr))
        ws.cell(row=(list_index[i]+list_index[i+1])/2, column=n_header+r).alignment=Alignment(
            horizontal='right',
            vertical='center'
        )
        n_roundr+=1
        temp.append((list_index[i]+list_index[i+1])/2)
        i+=2
        
    list_index=temp
    r+=1

ws.cell(row=list_index[0], column=n_header+r).border=Border(
    outline=True,
    bottom=Side('medium', color='FF000000')
)
wb.save('tournament.xlsx')